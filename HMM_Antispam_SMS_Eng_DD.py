# -*- coding: utf-8 -*-
"""
Created on Tue May 12 11:17:19 2020

@author: XiaTian
"""


from pyhanlp import *
import numpy as np
import openpyxl
import matplotlib.pyplot as pyplot 
from pomegranate import State, DiscreteDistribution, MultivariateGaussianDistribution, HiddenMarkovModel, GeneralMixtureModel, NormalDistribution, ExponentialDistribution
from sklearn.neural_network import MLPClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import hinge_loss, log_loss
from collections import Counter

def loadDataSet(FileName):
    Spam_Doc=[]
    White_Doc=[]
    f = open(FileName,'r', encoding='UTF-8')
    lines = f.readlines()
    for line in lines:
        if str.strip(line)!="":
            s=line.split('\t')
            Label=s[0]
            Doc = s[1]
            if Label=="spam":
                Spam_Doc.append(Doc)
            if Label=="ham":
                White_Doc.append(Doc)
    f.close()
    return White_Doc,Spam_Doc

def loadExcelDataSet(FileName):
    Spam_Doc=[]
    White_Doc=[]
    # Open excel file
    wb = openpyxl.load_workbook(FileName)
    ws = wb.active  
    col1=ws['A']
    col2=ws['B']
    i=0
    for row in col1:
        if col1[i].value=="spam":
            Spam_Doc.append(col2[i].value)
        if col1[i].value=="ham":
            White_Doc.append(col2[i].value)
        i=i+1
    return White_Doc,Spam_Doc

def createDocSegment(dataSet):
    segDoc=[]
    for sentence in dataSet: 
        #print(sentence)
        segDoc.append(HanLP.segment(sentence))
    return segDoc

def replaceAll(input, toReplace, replaceWith ): #
    while ( input.find(toReplace ) > -1 ):
        input = input.replace(toReplace, replaceWith )  
    return input

def OnlyGetTerm(document):
    CoreStopWordDictionary = JClass("com.hankcs.hanlp.dictionary.stopword.CoreStopWordDictionary")
    CoreStopWordDictionary.apply(document)
    TermSetinDoc=[]
    BlockedWordSet=[]#["-", "_","<",">",">>","<<","#","><","<>","#&gt","&lt","\'\'","\'","&amp",'&', '&Cs', '&SAM', '&XXX', '&first', '&gt', '&it', '&othrs']
    for term in document:
        StripTerm=str.strip(term.word)# Get word，its nature is term.nature
        '''
        StripTerm=replaceAll(StripTerm, "\'","")
        StripTerm=replaceAll(StripTerm, '\"','')
        StripTerm=replaceAll(StripTerm, "<","")
        StripTerm=replaceAll(StripTerm, ">","")
        StripTerm=replaceAll(StripTerm, "-","")
        StripTerm=replaceAll(StripTerm, "@","")
        '''
        if StripTerm!="":
            if not (term.nature.startsWith('a') or (term.word in BlockedWordSet) or (term.word.isalpha() and len(term.word)<=1)):  
                TermSetinDoc.append(StripTerm) 
                #if term.word=="-":
                #    print(term.word, term.nature, end="\t")
    return TermSetinDoc

def createVocabList(dataSet):
#get a set of all words
    vocabSet = set([])  
    for document in dataSet:
        TermSetinDoc = OnlyGetTerm(document)
        vocabSet = vocabSet | set(TermSetinDoc) # get union
    return list(vocabSet)

def  UniteSet(Seg_Doc_Set, Seg_Spam_set):
    Unite_Set=[]
    for doc in Seg_Doc_Set:
        Unite_Set.append(doc)
    for doc in Seg_Spam_set:
        Unite_Set.append(doc)    
    return Unite_Set

def FilterDoc(Doc_Set):
    Filtered_Doc_set=[]
    for doc in Doc_Set:
        Filtered_Doc_set.append(OnlyGetTerm(doc))
    return Filtered_Doc_set
    
if __name__ == "__main__":
    #---------------read white sms
    print("Load Data and Split Words ",end=".............\n")
    Doc_Set,Spam_Set = loadDataSet("SMSSpamCollection.txt")
    #Doc_Set,Spam_Set = loadExcelDataSet("SMSSpamCollection.xlsx")
    #Doc_Set=Doc_Set[0:746] #enable this line for balanced dataset test
    Seg_Doc_Set = createDocSegment(Doc_Set)
    Seg_Spam_set= createDocSegment(Spam_Set)
    
    #gen union term set of all doc
    print("gen union term set of all doc",end=".............\n")
    Term_Set = createVocabList(list(set(Seg_Doc_Set).union(set(Seg_Spam_set))))
    Term_Set.sort()
    #print (Term_Set)
    print ("Seg_Doc_Set len:",len(Seg_Doc_Set))
    print ("Seg_Spam_set len:",len(Seg_Spam_set))
    print ("Term_Set len:",len(Term_Set))
    
    #Gen Doc vec
    Filtered_Seg_Doc_Set=FilterDoc(Seg_Doc_Set)
    Filtered_Seg_Spam_Set=FilterDoc(Seg_Spam_set)
 
    #Train set
    Observation_train = []
    Observation_ham_train = []
    Observation_spam_train = []
    #Observation_White_train=[]
    #Observation_Spam_train=[]
    Labels_train = []
    for doc in Filtered_Seg_Doc_Set:
        #Observation_White_train.extend(doc)
        Observation_train.extend(np.array(doc))
        Observation_ham_train.extend(np.array(doc))
        Labels_train.extend(["-White-"] * len(doc))
    for doc in Filtered_Seg_Spam_Set:
        #Observation_Spam_train.extend(doc)
        Observation_train.extend(np.array(doc))
        Observation_spam_train.extend(np.array(doc))
        Labels_train.extend(["-Spam-"] * len(doc))
    #print(Observation_train, "\n",Labels_train)
        
    #Discrete Distribution HMM
    Dict={}
    for term in Term_Set:
        Dict[term]=0
    #print(len(Dict),len(Term_Set))
    D_ham = DiscreteDistribution(Dict)
    D_ham.fit(Observation_ham_train)


    #Discrete Distribution HMM
    Dict={}
    for term in Term_Set:
        Dict[term]=0
    print(len(Dict),len(Term_Set))
    D_spam = DiscreteDistribution(Dict)
    D_spam.fit(Observation_spam_train)
    
    #dump data for ploting
    doc1=open('output1.txt','w')
    for item in D_ham.items():
        print('{:.20f}'.format(item[1]),file=doc1)
    doc2=open('output2.txt','w')
    for item in D_spam.items():
        print('{:.20f}'.format(item[1]),file=doc2)
        
    print("Train Discrete HMM",end=".............\n")
    
    s1=State(D_ham, name="-White-")
    s2=State(D_spam, name="-Spam-")
    
    model = HiddenMarkovModel()
    model.add_states(s1, s2)
    model.add_transition(model.start, s1, 0.5)
    model.add_transition(model.start, s2, 0.5)
    model.add_transition(model.end, s1, 0.5)
    model.add_transition(model.end, s2, 0.5)
    model.add_transition(s1, s1, 0.5)
    model.add_transition(s1, s2, 0.5)
    model.add_transition(s2, s1, 0.5)
    model.add_transition(s2, s2, 0.5)
    model.bake()
    
    model.fit([Observation_train],labels=[Labels_train],algorithm = 'labeled')
    
    #print(model)
    #model.plot()
    print("done",end="\n")
    for state in model.states:
        print("hmm_state:", state.name)
    
    #print("\n\nTest for White SMS(state 1 is White):")
    Standard_HMM_correct=0
    Standard_HMM_wrong=0
    White_Correct=0
    White_Wrong=0
    Spam_Correct=0
    Spam_Wrong=0
    for doc in Filtered_Seg_Doc_Set:
        Y_pred=model.predict(doc[:(len(doc)-1)],algorithm='viterbi')
        #print(doc)
        #print("viterbi",Y_pred,end="\n")
        Dicision=max(Y_pred, key=Y_pred.count) 
        if len(Y_pred)==2:
            Dicision=Y_pred[1]
        if Dicision==1:
            Standard_HMM_correct=Standard_HMM_correct+1
            White_Correct=White_Correct+1
        else:
            Standard_HMM_wrong=Standard_HMM_wrong+1
            White_Wrong=White_Wrong+1
            #print(doc)
            #print("viterbi",Y_pred,end="\n")
    #print("\n\nTest for Spam SMS(state 0 is Spam):")
    for doc in Filtered_Seg_Spam_Set:
        Y_pred=model.predict(doc[:(len(doc)-1)],algorithm='viterbi')
        #print(doc)
        #print("viterbi",Y_pred,end="\n")
        Dicision=max(Y_pred, key=Y_pred.count) 
        if len(Y_pred)==2:
            Dicision=Y_pred[1]
        if Dicision==0:
            Standard_HMM_correct=Standard_HMM_correct+1
            Spam_Correct=Spam_Correct+1
        else:
            Standard_HMM_wrong=Standard_HMM_wrong+1
            Spam_Wrong=Spam_Wrong+1
            #print(doc)
            #print("viterbi",Y_pred,end="\n")
    Standard_HMM_Accuracy=Standard_HMM_correct/(Standard_HMM_correct+Standard_HMM_wrong)
    print("Summary:")
    print("Total Spam SMS:", len(Spam_Set),"Total White SMS:",len(Doc_Set))
    print("Standard HMM Result: \nTotal Correct：", Standard_HMM_correct)
    print("Total Wrong：", Standard_HMM_wrong)
    TP=White_Correct
    FP=White_Wrong
    TN=Spam_Correct
    FN=Spam_Wrong
    White_SMS_Precision=TP/(TP+FP)
    White_SMS_Recall=TP/(TP+FN)
    White_SMS_F1=2*White_SMS_Precision*White_SMS_Recall/(White_SMS_Precision+White_SMS_Recall)
    Spam_SMS_Precision=TN/(TN+FN)
    Spam_SMS_Recall=TN/(TN+FP)
    Spam_SMS_F1=2*Spam_SMS_Precision*Spam_SMS_Recall/(Spam_SMS_Precision+Spam_SMS_Recall)
    print('In detail:')
    print('White_Correct:',White_Correct,'White_Wrong:',White_Wrong)
    print('White SMS Precision：{:.3f}'.format(White_SMS_Precision))    
    print('White SMS Recall: {:.3f}'.format(White_SMS_Recall))    
    print('White SMS F1: {:.3f}'.format(White_SMS_F1))    

    print('Spam_Correct:',Spam_Correct,'Spam_Wrong:',Spam_Wrong)
    print('SPAM SMS Precision：{:.3f}'.format(Spam_SMS_Precision))    
    print('SPAM SMS Recall:{:.3f}'.format(Spam_SMS_Recall))    
    print('SPAM SMS F1:{:.3f}'.format(Spam_SMS_F1))   

    print('The discrete HMM Accuracy：{:.3f}'.format((TN+TP)/(TN+TP+FN+FP)))
    AUC= (TP/(TP+FN)+(TN/(TN+FP)))/2
    print('AUC:{:.3f}'.format(AUC))  
    #Draw Figures
    #Term Frequency Figure
    MostOccurWords=[]
    count_Spam_Term_Frequency=[1]*len(Term_Set)
    for doc in Filtered_Seg_Spam_Set:
        for term in set(doc):
            TF=doc.count(term)
            if TF>1:
                if count_Spam_Term_Frequency[Term_Set.index(term)]<TF:
                    count_Spam_Term_Frequency[Term_Set.index(term)]=TF
                    MostOccurWords_set = set(MostOccurWords)
                    if TF>=4 and not term in MostOccurWords_set:
                        MostOccurWords.append(term)
    count_SpamTerm=0
    count_White_Term_Frequency=[1]*len(Term_Set)
    for doc in Filtered_Seg_Doc_Set:
         for term in set(doc):
            TF=doc.count(term)
            if TF>1:
                if count_White_Term_Frequency[Term_Set.index(term)]<TF:
                    count_White_Term_Frequency[Term_Set.index(term)]=TF
                    MostOccurWords_set = set(MostOccurWords)
                    if TF>=4 and not term in MostOccurWords_set:
                        MostOccurWords.append(term)
    print("Term Summary:\nTotal Terms:",len(Term_Set))
    print("Spam Term Frequency Counter:",Counter(count_Spam_Term_Frequency))
    print("White Term Frequency Counter:",Counter(count_White_Term_Frequency))
    print("Words with high frequency", MostOccurWords)